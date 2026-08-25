# backend/api/routes_export.py
# ProspectLens — Export Endpoints
#
# Generates downloadable CSV and XLSX files from collected leads.
#
# Endpoints:
#   POST /api/export/csv    — Export leads to CSV
#   POST /api/export/xlsx   — Export leads to XLSX (Excel)
#   GET  /api/export/history — View all past exports

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlmodel import Session, select
from pydantic import BaseModel
from typing import Optional
from pathlib import Path
from datetime import datetime
import csv
import os

from database.db import get_session
from database.models import Lead, Contact, ExportHistory

router = APIRouter(tags=["Export"])

# Exports are saved in a writeable user-data directory
local_app_data = os.getenv("LOCALAPPDATA")
if local_app_data:
    EXPORTS_DIR = Path(local_app_data) / "ProspectLens" / "exports"
else:
    EXPORTS_DIR = Path.home() / "Downloads" / "ProspectLens"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ==============================================================================
# REQUEST SCHEMA
# ==============================================================================

class ExportRequest(BaseModel):
    batch_id:     Optional[str] = None     # Export specific batch, or all if None
    source_site:  Optional[str] = None     # Filter by platform
    lead_status:  Optional[str] = None     # Filter by status
    format:       str = "xlsx"             # csv or xlsx


# ==============================================================================
# HELPER — Fetch leads for export with optional filters
# ==============================================================================

def _fetch_leads_for_export(
    session: Session,
    batch_id: Optional[str],
    source_site: Optional[str],
    lead_status: Optional[str]
) -> list:
    statement = select(Lead)
    if batch_id:
        statement = statement.where(Lead.batch_id == batch_id)
    if source_site:
        statement = statement.where(Lead.source_site == source_site)
    if lead_status:
        statement = statement.where(Lead.lead_status == lead_status)
    statement = statement.order_by(Lead.collected_at.desc())
    return session.exec(statement).all()


# ==============================================================================
# HELPER — Get first phone and email for a lead
# ==============================================================================

def _get_primary_contacts(lead_id: str, session: Session) -> dict:
    contacts = session.exec(
        select(Contact).where(Contact.lead_id == lead_id)
    ).all()

    phone    = next((c.contact_value for c in contacts if c.contact_type == "phone"),    "")
    email    = next((c.contact_value for c in contacts if c.contact_type == "email"),    "")
    whatsapp = next((c.contact_value for c in contacts if c.contact_type == "whatsapp"), "")

    return {"phone": phone, "email": email, "whatsapp": whatsapp}


# ==============================================================================
# HELPER — Build flat row dict for each lead (used in both CSV and XLSX)
# ==============================================================================

def _build_export_rows(leads: list, session: Session) -> list[dict]:
    rows = []
    for lead in leads:
        contacts = _get_primary_contacts(lead.lead_id, session)
        rows.append({
            "Business Name":   lead.business_name,
            "Service":         lead.service_name    or "",
            "Contact Person":  lead.contact_person  or "",
            "Phone":           contacts["phone"],
            "WhatsApp":        contacts["whatsapp"],
            "Email":           contacts["email"],
            "Website":         lead.website         or "",
            "Address":         lead.address         or "",
            "City":            lead.city            or "",
            "State":           lead.state           or "",
            "Postal Code":     lead.postal_code     or "",
            "Category":        lead.category        or "",
            "Source":          lead.source_site,
            "Listing URL":     lead.listing_url     or "",
            "Lead Status":     lead.lead_status,
            "Notes":           lead.notes           or "",
            "Tags":            lead.tags            or "",
            "Collected At":    lead.collected_at.strftime("%Y-%m-%d %H:%M") if lead.collected_at else "",
            
            # New fields for Sprint 4.3
            "Search Keyword":  lead.search_keyword  or "",
            "Search Location": lead.search_location or "",
            "Search Query":    lead.search_query    or "",
            "Search URL":      lead.directory_search_url or "",
            "Collection Date": lead.collection_date or "",
            "Collection Time": lead.collection_time or "",
            "Rating":          lead.rating          if lead.rating is not None else "",
            "Review Count":    lead.review_count    if lead.review_count is not None else "",
            "Open Status":     lead.open_status     or "",
            "Price Level":     lead.price_level     or "",
            "Displayed Price": lead.displayed_price or "",
            "Price Type":      lead.price_type      or "",
            "Website Domain":  lead.website_domain  or "",
            "Flexible Metadata": lead.flexible_metadata or "",
            
            # Sprint 4.4 Universal Schema additions
            "Sub Category":    lead.sub_category    or "",
            "Source Business ID": lead.source_business_id or "",
            "Collector Version": lead.collector_version or "1.0.0",
            "Secondary Phones": lead.secondary_phones or "",
        })
    return rows


# ==============================================================================
# POST /api/export/csv
# Generates a CSV file and returns it as a download.
# ==============================================================================

@router.post("/export/csv")
def export_csv(req: ExportRequest, session: Session = Depends(get_session)):
    leads = _fetch_leads_for_export(session, req.batch_id, req.source_site, req.lead_status)

    if not leads:
        raise HTTPException(status_code=404, detail="No leads found matching the filters")

    rows     = _build_export_rows(leads, session)
    filename = f"prospectlens_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = EXPORTS_DIR / filename

    # Write CSV
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        # utf-8-sig adds BOM so Excel opens it correctly with Indian characters
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    # Log export to history
    file_size_kb = round(os.path.getsize(filepath) / 1024, 2)
    export_log = ExportHistory(
        batch_id=req.batch_id,
        export_format="csv",
        file_path=str(filepath),
        record_count=len(rows),
        file_size_kb=file_size_kb
    )
    session.add(export_log)
    session.commit()

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="text/csv"
    )


# ==============================================================================
# POST /api/export/xlsx
# Generates an Excel XLSX file with formatted headers and returns it as download.
# ==============================================================================

@router.post("/export/xlsx")
def export_xlsx(req: ExportRequest, session: Session = Depends(get_session)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    leads = _fetch_leads_for_export(session, req.batch_id, req.source_site, req.lead_status)

    if not leads:
        raise HTTPException(status_code=404, detail="No leads found matching the filters")

    rows     = _build_export_rows(leads, session)
    filename = f"prospectlens_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = EXPORTS_DIR / filename

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ProspectLens Leads"

    # Header style — dark green background, white bold text
    header_fill = PatternFill(start_color="1A3C1A", end_color="1A3C1A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = list(rows[0].keys())

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row.values(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    wb.save(filepath)

    # Log export to history
    file_size_kb = round(os.path.getsize(filepath) / 1024, 2)
    export_log = ExportHistory(
        batch_id=req.batch_id,
        export_format="xlsx",
        file_path=str(filepath),
        record_count=len(rows),
        file_size_kb=file_size_kb
    )
    session.add(export_log)
    session.commit()

    return FileResponse(
        path=filepath,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ==============================================================================
# GET /api/export/history
# View all past exports — useful to re-download previous files.
# ==============================================================================

@router.get("/export/history")
def export_history(session: Session = Depends(get_session)):
    exports = session.exec(
        select(ExportHistory).order_by(ExportHistory.exported_at.desc())
    ).all()

    return {
        "status": "ok",
        "count": len(exports),
        "exports": exports
    }